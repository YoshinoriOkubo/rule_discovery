import calendar as cal
import numpy as np
import datetime
import os
import openpyxl
import matplotlib.pyplot as plt
from constants import *

def convert2to10_in_list(list):
    result = 0
    length = len(list)
    for i in range(len(list)):
        x = length - 1 - i
        result += list[i] * 2 ** (x)
    if len(list) == 4:
        return GRAY_CODE_4[result]
    elif len(list) == 1:
        return result
    else:
        return None

def export_population(list):
    path = '../output/ship_rule.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for index in range(len(list)):
        individual = list[index]
        row = index + 2
        for col_cond in range(DEFAULT_NUM_OF_CONDITION*2):
            if col_cond == 0 or col_cond == 1:
                sheet.cell(row = row, column = col_cond + 1).value = OIL_PRICE_LIST[convert2to10_in_list(individual[col_cond])]
            elif col_cond == 2 or col_cond == 3:
                sheet.cell(row = row, column = col_cond + 1).value = FREIGHT_RATE_LIST[convert2to10_in_list(individual[col_cond])]
            elif col_cond == 4 or col_cond == 5:
                sheet.cell(row = row, column = col_cond + 1).value = EXCHANGE_RATE_LIST[convert2to10_in_list(individual[col_cond])]
            else:
                sheet.cell(row = row, column = col_cond + 1).value = OWN_SHIP_LIST[convert2to10_in_list(individual[col_cond])]
        for col_act in range(DEFAULT_NUM_OF_ACTION):
            sheet.cell(row = row, column = col_act + DEFAULT_NUM_OF_CONDITION*2 + 1).value = individual[DEFAULT_NUM_OF_CONDITION*2+col_act]
        sheet.cell(row = row, column = DEFAULT_NUM_OF_CONDITION*2 + DEFAULT_NUM_OF_ACTION + 1).value = individual[-1][0]
        sheet.cell(row = row, column = DEFAULT_NUM_OF_CONDITION*2 + DEFAULT_NUM_OF_ACTION + 2).value = individual[-1][1]
    wb.save(path)
    wb.close()
    print('saving changes')


def load_generated_sinario():
    all_data = []
    for i in range(4):
        if i == 0:
            history_data_path = '../output/oil_price.csv'
        elif i == 1:
            history_data_path = '../output/freight_outward.csv'
        elif i == 2:
            history_data_path = '../output/freight_return.csv'
        elif i == 3:
            history_data_path = '../output/exchange_rate.csv'


        # read data
        dt   = np.dtype({'names': ('date', 'price'),
                       'formats': ('S10' , np.float)})
        data = np.array([], dtype=dt)
        for j in range(DEFAULT_PREDICT_PATTERN_NUMBER):
            data = np.append(data,np.genfromtxt(history_data_path,
                             delimiter=',',
                             dtype=dt,
                             usecols=[2*j,2*j+1],
                             skip_header=0,
                             encoding='utf-8_sig'))
        data = data.reshape(DEFAULT_PREDICT_PATTERN_NUMBER,VESSEL_LIFE_TIME*12)
        all_data.append(data)
    return all_data

#load history data of crude oil
def load_monthly_history_data(from_date=None, to_date=None):
    history_data_path = '../data/crude_oil_monthly_history.csv'
    # read data
    dt   = np.dtype({'names': ('date', 'price'),
                   'formats': ('S10' , np.float)})
    data = np.genfromtxt(history_data_path,
                         delimiter=',',
                         dtype=dt,
                         usecols=[0,1],
                         skip_header=1)

    # from_date
    if not from_date is None:
        from_datetime = datetime.datetime.strptime(from_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if from_datetime <= datetime_key:
                break
        data = data[index:]

    # to_date
    if not to_date is None:
        to_datetime = datetime.datetime.strptime(to_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if to_datetime <= datetime_key:
                break
        data = data[:index]

    return data

#load history data of freight rate
def load_monthly_freight_rate_data(direction,from_date=None, to_date=None):
    if direction == OUTWARD:
        history_data_path = '../data/freight_rate_history_outward.csv'
    else:
        if direction == RETURN:
            history_data_path = '../data/freight_rate_history_return.csv'
        else:
            if direction == CCFI:
                history_data_path = '../data/ccfi_history.csv'
            else:
                raise Exception('Error!')
    # read data
    dt   = np.dtype({'names': ('date', 'price'),
                   'formats': ('S10' , np.float)})
    data = np.genfromtxt(history_data_path,
                         delimiter=',',
                         dtype=dt,
                         usecols=[0,1],
                         skip_header=1)

    # from_date
    if not from_date is None:
        from_datetime = datetime.datetime.strptime(from_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if from_datetime <= datetime_key:
                break
        data = data[index:]

    # to_date
    if not to_date is None:
        to_datetime = datetime.datetime.strptime(to_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if to_datetime <= datetime_key:
                break
        data = data[:index]

    return data

#load history data of freight rate
def load_monthly_exchange_rate_data(from_date=None, to_date=None):
    history_data_path = '../data/exchange_rate_after_praza_agreement.csv'

    # read data
    dt   = np.dtype({'names': ('date', 'price'),
                   'formats': ('S10' , np.float)})
    data = np.genfromtxt(history_data_path,
                         delimiter=',',
                         dtype=dt,
                         usecols=[0,1],
                         skip_header=1)

    # from_date
    if not from_date is None:
        from_datetime = datetime.datetime.strptime(from_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if from_datetime <= datetime_key:
                break
        data = data[index:]

    # to_date
    if not to_date is None:
        to_datetime = datetime.datetime.strptime(to_date, "%Y/%m/%d")
        for index in range(len(data)):
            datetime_key = datetime.datetime.strptime(data['date'][index], "%Y/%m/%d")
            if to_datetime <= datetime_key:
                break
        data = data[:index]

    return data

# substitute inf to nan in values
def inf_to_nan_in_array(values):
    inf_induces = np.where(values==float('-inf'))[0]

    for i in range(len(inf_induces)):
        values[inf_induces[i]] = float('nan')

    return values

def add_month(current_date, num=1):
    for _n in range(num):
        _d, days      = cal.monthrange(current_date.year, current_date.month)
        current_date += datetime.timedelta(days=days)
    return current_date

def add_year(start_date, year_num=1):
    current_date = start_date
    if year_num < 1:
        current_date = add_month(current_date, int(year_num * 12))
        return current_date

    for year_index in range(year_num):
        current_date = add_month(current_date, 12)
    return current_date

def prob(prob_value):
    N = 10000
    nonzero_num = np.count_nonzero(np.random.binomial(1, prob_value, N))
    threshold   = N * prob_value
    return (nonzero_num > threshold)

def calc_statistics(list):
    n = len(list)
    e = 0
    sigma = 0
    for i in range(n):
        e += list[i]
    e /= n
    for i in range(n):
        sigma += (list[i] - e)**2
    sigma /= n
    return [e,sigma]

def depict_distribution(oil,freight_outward,exchange):
    distribution = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    for pattern in range(DEFAULT_PREDICT_PATTERN_NUMBER):
        for a in range(VESSEL_LIFE_TIME * 12):
            f = oil[pattern][a]['price']
            for i in range(16):
                if i == 15:
                    if OIL_PRICE_LIST[i] < f:
                        distribution[i] += 1
                else:
                    if OIL_PRICE_LIST[i] < f and f < OIL_PRICE_LIST[i+1]:
                        distribution[i] += 1
    for index in range(len(distribution)):
        distribution[index] = distribution[index]/(DEFAULT_PREDICT_PATTERN_NUMBER*180.0)
    left = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    label = OIL_PRICE_LIST
    plt.title('Oil price distribution')
    plt.xlabel('oil price')
    plt.ylabel('Propability')
    plt.bar(left,distribution,tick_label=label,align='center')
    save_dir = '../output'
    plt.savefig(os.path.join(save_dir, 'oil_price_distribution.png'))
    plt.close()

    distribution = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    for pattern in range(DEFAULT_PREDICT_PATTERN_NUMBER):
        for x in range(VESSEL_LIFE_TIME*12):
            f = freight_outward[pattern][x]['price']
            for i in range(16):
                if i ==15:
                    if FREIGHT_RATE_LIST[i] < f:
                        distribution[i] += 1
                else:
                    if FREIGHT_RATE_LIST[i] < f and f < FREIGHT_RATE_LIST[i+1]:
                        distribution[i] += 1
        _x = range(0,VESSEL_LIFE_TIME*12)
    for index in range(len(distribution)):
        distribution[index] = distribution[index]/(DEFAULT_PREDICT_PATTERN_NUMBER*180.0)
    #left = [200,300,400,500,600,700,800,900]
    left = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    label = FREIGHT_RATE_LIST
    plt.title('Freight distribution')
    plt.xlabel('freight outward')
    plt.ylabel('Propability')
    plt.bar(left,distribution,tick_label=label,align='center')
    save_dir = '../output'
    plt.savefig(os.path.join(save_dir, 'freight_distribution.png'))
    plt.close()

    distribution = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    for pattern in range(DEFAULT_PREDICT_PATTERN_NUMBER):
        for x in range(VESSEL_LIFE_TIME*12):
            f = exchange[pattern][x]['price']
            for i in range(16):
                if i ==15:
                    if EXCHANGE_RATE_LIST[i] < f:
                        distribution[i] += 1
                else:
                    if EXCHANGE_RATE_LIST[i] < f and f < EXCHANGE_RATE_LIST[i+1]:
                        distribution[i] += 1
        _x = range(0,VESSEL_LIFE_TIME*12)
    for index in range(len(distribution)):
        distribution[index] = distribution[index]/(DEFAULT_PREDICT_PATTERN_NUMBER*180.0)
    #left = [200,300,400,500,600,700,800,900]
    left = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
    label = EXCHANGE_RATE_LIST
    plt.title('Exchange distribution')
    plt.xlabel('exchange')
    plt.ylabel('Propability')
    plt.bar(left,distribution,tick_label=label,align='center')
    save_dir = '../output'
    plt.savefig(os.path.join(save_dir, 'exchange_distribution.png'))
    plt.close()
