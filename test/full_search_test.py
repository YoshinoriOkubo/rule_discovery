import numpy as np
import sys
sys.path.append('../models')
from oil_price import Sinario
from freight_rate_outward import FreightOutward
from freight_rate_return import FreightReturn
from ship import Ship
import matplotlib.pyplot as plt
from ga import GA
# import own modules #
sys.path.append('../public')
from my_modules import *
from constants  import *
import openpyxl
import math

path='../output/oil_price.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook['Sheet1']
oil = []
for i in range(180):
    oil.append(sheet.cell(row = i + 1, column = 1).value)
workbook.close()

path='../output/freight_outward.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook['Sheet1']
freight_outward = []
for i in range(180):
    freight_outward.append(sheet.cell(row = i + 1, column = 1).value)
workbook.close()

path='../output/freight_return.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook['Sheet1']
freight_return = []
for i in range(180):
    freight_return.append(sheet.cell(row = i + 1, column = 1).value)
workbook.close()

Ship = ship = Ship(TEU_SIZE,INITIAL_SPEED,ROUTE_DISTANCE)
cash = 0
charter = []

a = 6
b = 100
c = 145
d = 132
e = 168
charter_timing = [a,b,c,d,e]
for i in charter_timing:
    current_oil_price = oil[i]
    current_freight_rate_outward = freight_outward[i]
    current_freight_rate_return = freight_return[i]
    total_freight = 0.5 * ( current_freight_rate_outward * LOAD_FACTOR_ASIA_TO_EUROPE + current_freight_rate_return * LOAD_FACTOR_EUROPE_TO_ASIA)
    charter.append(ship.calculate_income_per_month(current_oil_price,total_freight)*RISK_PREMIUM)
x = 0
for year in range(15):
    cash_year = 0
    for month in range(12):
        #current_oil_price = oil[year*12+month]
        #current_freight_rate_outward = freight_outward[year*12+month]
        #current_freight_rate_return = freight_return[year*12+month]
        #total_freight = 0.5 * ( current_freight_rate_outward * LOAD_FACTOR_ASIA_TO_EUROPE + current_freight_rate_return * LOAD_FACTOR_EUROPE_TO_ASIA)
        #cash_year += ship.calculate_income_per_month(current_oil_price,total_freight)
        if year*12+month >= a and year*12+month < a + 36:
            cash_year += charter[0]
        elif year*12+month >= b and year*12+month < b + 36:
            cash_year += charter[1]
        elif year*12+month >= c and year*12+month < c + 36:
            cash_year += charter[2]
        #elif year*12+month >= d and year*12+month < d + 36:
        #    cash_year += charter[3]
        #elif year*12+month >= e and year*12+month < e + 36:
        #    cash_year += charter[4]
        else:
            current_oil_price = oil[year*12+month]
            current_freight_rate_outward = freight_outward[year*12+month]
            current_freight_rate_return = freight_return[year*12+month]
            total_freight = 0.5 * ( current_freight_rate_outward * LOAD_FACTOR_ASIA_TO_EUROPE + current_freight_rate_return * LOAD_FACTOR_EUROPE_TO_ASIA)
            cash_year += ship.calculate_income_per_month(current_oil_price,total_freight)
    if year < DEPRECIATION_TIME:
        cash_year -= INITIAL_COST_OF_SHIPBUIDING/DEPRECIATION_TIME
    DISCOUNT = (1 + DISCOUNT_RATE) ** (year + 1)
    cash += cash_year/(DISCOUNT)
print(cash/HUNDRED_MILLION)
