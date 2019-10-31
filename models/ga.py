import random
import copy
import time
from tqdm import tqdm
import sys
import matplotlib.pyplot as plt
import os
import openpyxl
from multiprocessing import Pool
import multiprocessing as multi
from ship import Ship
from oil_price import Sinario
from freight_rate_outward import FreightOutward
from freight_rate_return import FreightReturn
from exchange_rate import ExchangeRate
# import own modules #
sys.path.append('../public')
sys.path.append('../output')
from constants  import *
from my_modules import *

class GA:

    def __init__(self,oil_price_data,freight_rate_outward,freight_rate_return,exchange_rate,TEU_size,init_speed,route_distance,actionlist,generation=None,population_size=None,alpha=None,crossover_rate=None):
        self.oil_price_data = oil_price_data #oil_price_history_data
        self.freight_rate_outward_data = freight_rate_outward #feright rate outward history data
        self.freight_rate_return_data = freight_rate_return # freight rate return history data
        self.exchange_rate_data = exchange_rate # exchange_rate history data
        self.TEU_size = TEU_size #size of ship(TEU)
        self.init_speed = init_speed # initial speed of ship (km/h)
        self.route_distance = route_distance # distance of fixed route (km)
        self.actionlist = actionlist # decision of action parts.
        self.generation = generation if generation else DEFAULT_GENERATION # the number of generation
        self.population_size = population_size if population_size else DEFAULT_POPULATION_SIZE  # the number of individual
        self.alpha = alpha if alpha else DEFAULT_ALPHA # the rate of mutation
        self.crossover_rate = crossover_rate if crossover_rate else DEFAULT_CROSSOVER_RATE
        self.priority = []
        self.population = [] # population that has individual
        self.temp = [] #temporary group that has individuals
        self.bestpopulation = [] # group that has the best individuals in each generation
        self.averagepopulation = [] # the average value of fitness in each generation
        self.compare_rule = []
        if self.decision == DECISION_INTEGRATE:
            for rule_index in range(len(RULE_SET)):
                self.compare_rule.append([])
                for i in range(DEFAULT_NUM_OF_CONDITION*2):
                    self.compare_rule[rule_index].append([0,0,0,0])
                if RULE_SET[rule_index] == DECISION_SPEED:
                    self.compare_rule[rule_index].append([1,1,0,1]) #19knot
                elif RULE_SET[rule_index] == DECISION_SELL or RULE_SET[rule_index] == DECISION_BUY:
                    self.compare_rule[rule_index].append(ACTION_STAY)
                elif RULE_SET[rule_index] == DECISION_CHARTER_OUT or RULE_SET[rule_index] == DECISION_CHARTER_IN:
                    self.compare_rule[rule_index].append([0,0])#charter period
                    self.compare_rule[rule_index].append(ACTION_STAY)
        else:
            for i in range(DEFAULT_NUM_OF_CONDITION*2):
                self.compare_rule.append([0,0,0,0])
            if self.decision == DECISION_SPEED:
                self.compare_rule.append([1,1,0,1]) #19knot
            elif self.decision == DECISION_SELL or self.decision == DECISION_BUY:
                self.compare_rule.append(ACTION_STAY)# ships number
            elif self.decision == DECISION_CHARTER_OUT or self.decision == DECISION_CHARTER_IN:
                self.compare_rule.append([0,0])#charter period
                self.compare_rule.append(ACTION_STAY)#ships number
        self.compare_rule.append([0,0])# average profit and varianve

    def convert2to10_in_list(self,list):
        result = 0
        length = len(list)
        for i in range(len(list)):
            x = length - 1 - i
            result += list[i] * 2 ** (x)
        if len(list) == 4:
            return GRAY_CODE_4[result]
        elif len(list) == 2:
            return GRAY_CODE_2[result]
        elif len(list) == 1:
            return result
        else:
            return None

    def adapt_rule(self,oil_price,freight,exchange,rule):
        if self.decision == DECISION_INTEGRATE:
            result = []
            for rule_index in range(len(rule)-1):
                rule_for_X = rule[rule_index]
                result.append([False])
                a = OIL_PRICE_LIST[self.convert2to10_in_list(rule_for_X[0])]
                b = OIL_PRICE_LIST[self.convert2to10_in_list(rule_for_X[1])]
                if a == b or ( a <= oil_price and oil_price <= b):
                    c = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule_for_X[2])]
                    d = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule_for_X[3])]
                    if c == d or ( c <= freight and freight <= d):
                        e = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule_for_X[4])]
                        f = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule_for_X[5])]
                        if e == f or (e <= exchange and exchange <= f):
                            if RULE_SET[rule_index] == DECISION_SPEED:
                                result[rule_index][0] = True
                                result[rule_index].append(VESSEL_SPEED_LIST[self.convert2to10_in_list(rule_for_X[-1])])
                            elif RULE_SET[rule_index] == DECISION_SELL:
                                result[rule_index][0] = True
                                result[rule_index].append(SELL_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                            elif RULE_SET[rule_index] == DECISION_BUY:
                                result[rule_index][0] = True
                                result[rule_index].append(BUY_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                            elif (RULE_SET[rule_index] == DECISION_CHARTER_OUT or RULE_SET[rule_index] == DECISION_CHARTER_IN) and rule_for_X[-1][0] != ACTION_STAY:
                                result[rule_index][0] = True
                                result[rule_index].append(CHARTER_PERIOD[self.convert2to10_in_list(rule_for_X[-2])])
                                result[rule_index].append(CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
            return result
        else:
            a = OIL_PRICE_LIST[self.convert2to10_in_list(rule[0])]
            b = OIL_PRICE_LIST[self.convert2to10_in_list(rule[1])]
            if a == b or ( a <= oil_price and oil_price <= b):
                c = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule[2])]
                d = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule[3])]
                if c == d or ( c <= freight and freight <= d):
                    e = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule[4])]
                    f = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule[5])]
                    if e == f or (e <= exchange and exchange <= f):
                        if self.decision == DECISION_SPEED:
                            return [True,VESSEL_SPEED_LIST[self.convert2to10_in_list(rule[-2])]]
                        elif self.decision == DECISION_SELL:
                            return [True,SELL_NUMBER[self.convert2to10_in_list(rule[-2])]]
                        elif self.decision == DECISION_BUY:
                            return [True,BUY_NUMBER[self.convert2to10_in_list(rule[-2])]]
                        elif self.decision == DECISION_CHARTER_OUT or self.decision == DECISION_CHARTER_IN:
                            return [True,CHARTER_PERIOD[self.convert2to10_in_list(rule[-3])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule[-2])]]
                        else:
                            print('selected decision item does not exist')
                            sys.exit()
            return [False]

    def crossing(self,a,b,num_block=None):
        temp1 = []
        temp2 = []
        if num_block == None:
            if self.decision == DECISION_INTEGRATE:
                for rule_index in range(len(a)-1):
                    temp1.append([])
                    temp2.append([])
                    if RULE_SET[rule_index] == DECISION_SPEED:
                        num_block = DEFAULT_NUM_OF_CONDITION*2 + 1
                    elif RULE_SET[rule_index] == DECISION_SELL:
                        num_block = DEFAULT_NUM_OF_CONDITION*2 + 1
                    elif RULE_SET[rule_index] == DECISION_BUY:
                        num_block = DEFAULT_NUM_OF_CONDITION*2 + 1
                    elif RULE_SET[rule_index] == DECISION_CHARTER_OUT or RULE_SET[rule_index] == DECISION_CHARTER_IN:
                        num_block = DEFAULT_NUM_OF_CONDITION*2 + 2
                    crossing_block = random.randint(0,num_block-1)
                    for x in range(num_block):
                        if x == crossing_block:
                            temp1[rule_index].append([])
                            temp2[rule_index].append([])
                            length = len(a[rule_index][x]) - 1
                            crossing_point = random.randint(0,length)
                            for i in range(0,crossing_point):
                                temp1[rule_index][x].append(a[rule_index][x][i])
                                temp2[rule_index][x].append(b[rule_index][x][i])
                            for i in range(crossing_point,len(a[rule_index][x])):
                                temp1[rule_index][x].append(b[rule_index][x][i])
                                temp2[rule_index][x].append(a[rule_index][x][i])
                        else:
                            temp1[rule_index].append(a[rule_index][x])
                            temp2[rule_index].append(b[rule_index][x])
                    for x in range(num_block,len(a[rule_index])):
                        temp1[rule_index].append(a[rule_index][x])
                        temp2[rule_index].append(b[rule_index][x])
            else:
                print('function crossing needs one more argment')
                sys.exit()
        else:
            crossing_block = random.randint(0,num_block-1)
            for x in range(num_block):
                if x == crossing_block:
                    temp1.append([])
                    temp2.append([])
                    length = len(a[x]) - 1
                    crossing_point = random.randint(0,length)
                    for i in range(0,crossing_point):
                        temp1[x].append(a[x][i])
                        temp2[x].append(b[x][i])
                    for i in range(crossing_point,len(a[x])):
                        temp1[x].append(b[x][i])
                        temp2[x].append(a[x][i])
                else:
                    temp1.append(a[x])
                    temp2.append(b[x])
            for x in range(num_block,len(a)-1):
                temp1.append(a[x])
                temp2.append(b[x])
        temp1.append([0,0])
        temp2.append([0,0])
        return [temp1,temp2]

    def mutation(self,individual):
        if self.decision == DECISION_INTEGRATE:
            for rule_index in range(len(individual)-1):
                if random.random() < 1/(len(individual)-1):
                    rule_for_X = individual[rule_index]
                    mutation_block = random.randint(0,len(rule_for_X)-1)
                    length = len(rule_for_X[mutation_block]) - 1
                    point = random.randint(0,length)
                    rule_for_X[mutation_block][point] = (rule_for_X[mutation_block][point] + 1) % 2
        else:
            mutation_block = random.randint(0,len(individual)-2)
            length = len(individual[mutation_block]) - 1
            point = random.randint(0,length)
            individual[mutation_block][point] = (individual[mutation_block][point] + 1) % 2
        return individual

    def fitness_function(self,rule,priority=None):
        Record = []
        for pattern in range(DEFAULT_PREDICT_PATTERN_NUMBER):
            fitness = 0
            ship = Ship(self.TEU_size,self.init_speed,self.route_distance)
            for year in range(VESSEL_LIFE_TIME):
                cash_flow = 0
                for month in range(12):
                    current_oil_price = self.oil_price_data[pattern][year*12+month]['price']
                    current_freight_rate_outward = self.freight_rate_outward_data[pattern][year*12+month]['price']
                    current_freight_rate_return = self.freight_rate_return_data[pattern][year*12+month]['price']
                    total_freight = 0.5 * ( current_freight_rate_outward * LOAD_FACTOR_ASIA_TO_EUROPE + current_freight_rate_return * LOAD_FACTOR_EUROPE_TO_ASIA)
                    current_exchange = self.exchange_rate_data[pattern][year*12+month]['price']
                    #change by argment
                    result = self.adapt_rule(current_oil_price,current_freight_rate_outward,current_exchange,rule)
                    if self.decision == DECISION_SPEED:
                        if result[0]:
                            ship.change_speed(result[1])
                    elif self.decision == DECISION_SELL:
                        if result[0]:
                            cash_flow += ship.sell_ship(self.freight_rate_outward_data[pattern],year*12+month,result[1])
                    elif self.decision == DECISION_BUY:
                        if result[0]:
                            cash_flow += ship.buy_ship(self.freight_rate_outward_data[pattern],year*12+month,result[1])
                    elif self.decision == DECISION_CHARTER_OUT or self.decision == DECISION_CHARTER_IN:
                        if result[0] and result[2] > 0:
                            ship.charter_ship(current_oil_price,total_freight,result[2],result[1],self.decision)
                        if ship.charter_flag == True:
                            cash_flow += ship.charter()
                            ship.end_charter()
                    elif self.decision == DECISION_INTEGRATE:
                        if result[0][0] == True:
                            ship.change_speed(result[0][1])
                        if result[1][0] == True:
                            cash_flow += ship.sell_ship(self.freight_rate_outward_data[pattern],year*12+month,result[1][1])
                        if result[2][0] == True:
                            cash_flow += ship.buy_ship(self.freight_rate_outward_data[pattern],year*12+month,result[2][1])
                        if result[3][0] == True and result[3][2] > 0:
                            ship.charter_ship(current_oil_price,total_freight,result[3][2],result[3][1],DECISION_CHARTER_OUT)
                        if result[4][0] == True and result[4][2] > 0:
                            ship.charter_ship(current_oil_price,total_freight,result[4][2],result[4][1],DECISION_CHARTER_IN)
                        if ship.charter_flag == True:
                            cash_flow += ship.charter()
                            ship.end_charter()
                    cash_flow += ship.calculate_income_per_month(current_oil_price,total_freight)
                    cash_flow += ship.add_age()
                    ship.change_speed(self.init_speed)
                DISCOUNT = (1 + DISCOUNT_RATE) ** (year + 1)
                cash_flow *= self.exchange_rate_data[pattern][year*12+11]['price']
                fitness += cash_flow / DISCOUNT
            ship.sell_ship(self.freight_rate_outward_data[pattern],VESSEL_LIFE_TIME*12-1,ship.exist_number)
            fitness -= INITIAL_COST_OF_SHIPBUIDING*INITIAL_NUMBER_OF_SHIPS*self.exchange_rate_data[pattern][0]['price']
            fitness /= HUNDRED_MILLION
            fitness /= INITIAL_NUMBER_OF_SHIPS
            Record.append(fitness)
        e, sigma = calc_statistics(Record)
        return [e,sigma]

    def multiproces_fitness(self,i):
        e, sigma = self.fitness_function(self.temp[i],self.priority)
        return [i,[e,sigma]]

    def generateIndividual(self,actionlist):
        temp = []
        for condition in range(DEFAULT_NUM_OF_CONDITION*2):
            temp.append([])
            for a in range(4):
                temp[condition].append(random.randint(0,1))
        for action in range(self.num_action_part):
            temp.append(actionlist[action])
        temp.append([0,0])
        temp[-1][0],temp[-1][1] = self.fitness_function(temp)
        return temp

    def exchange_rule(self):
        for k in range(len(self.temp)):
            if OIL_PRICE_LIST[self.convert2to10_in_list(self.temp[k][0])] > OIL_PRICE_LIST[self.convert2to10_in_list(self.temp[k][1])]:
                self.temp[k][0],self.temp[k][1] = self.temp[k][1],self.temp[k][0]
            if FREIGHT_RATE_LIST[self.convert2to10_in_list(self.temp[k][2])] > FREIGHT_RATE_LIST[self.convert2to10_in_list(self.temp[k][3])]:
                self.temp[k][2],self.temp[k][3] = self.temp[k][3],self.temp[k][2]
            if EXCHANGE_RATE_LIST[self.convert2to10_in_list(self.temp[k][4])] > EXCHANGE_RATE_LIST[self.convert2to10_in_list(self.temp[k][5])]:
                self.temp[k][4],self.temp[k][5] = self.temp[k][5],self.temp[k][4]

    def change_population_size(self,time):
        if time < self.generation / 3.0:
            self.population_size = random.randint(DEFAULT_POPULATION_SIZE,int(DEFAULT_POPULATION_SIZE*4/3))
        elif time < self.generation * 2 /3.0:
            self.population_size = random.randint(int(DEFAULT_POPULATION_SIZE*2/3),DEFAULT_POPULATION_SIZE)
        else:
            self.population_size = random.randint(int(DEFAULT_POPULATION_SIZE*1/3),int(DEFAULT_POPULATION_SIZE*2/3))

    def depict_fitness(self):
        x = range(0,len(self.bestpopulation))
        y = []
        z = []
        for i in range(len(self.bestpopulation)):
            y.append(self.bestpopulation[i][-1][0])
            z.append(self.averagepopulation[i])
        plt.plot(x, y, marker='o',label='best')
        plt.plot(x, z, marker='x',label='average')
        plt.title('Transition of fitness', fontsize = 20)
        plt.xlabel('generation', fontsize = 16)
        plt.ylabel('fitness value', fontsize = 16)
        plt.tick_params(labelsize=14)
        plt.grid(True)
        plt.legend(loc = 'lower right')
        save_dir = '../output'
        plt.savefig(os.path.join(save_dir, 'fitness_{}.png'.format(name)))
        plt.close()

    def depict_average_variance(self,gene=None,list=None):
        x = []
        y = []
        for i in range(self.population_size):
            if gene == 0:
                x.append(list[i][-1][0])
                y.append(list[i][-1][1])
            else:
                x.append(self.population[i][-1][0])
                y.append(self.population[i][-1][1])
        plt.scatter(x,y)
        x_min = min(x)
        x_min = x_min*0.9 if x_min>0 else x_min*1.1
        plt.xlim(-1,1)
        plt.ylim(0,2)
        plt.title("Rule Performance")
        plt.xlabel("Expectation")
        plt.ylabel("Variance")
        plt.grid(True)
        save_dir = '../output'
        if gene == 0:
            plt.savefig(os.path.join(save_dir, 'Evaluation_{}_initial.png'.format(name)))
        else:
            plt.savefig(os.path.join(save_dir, 'Evaluation_{}.png'.format(name)))
        plt.close()

    def export_excel(self,initial=None):
        rule_type = ''
        if initial is None:
            path = '../output/ship_rule_{0}.xlsx'.format(rule_type)
        else:
            path = '../output/ship_rule_{0}_initial.xlsx'.format(rule_type)
        wb = openpyxl.load_workbook(path)
        sheet = wb['Sheet1']
        if self.decision == DECISION_INTEGRATE:
            for i in range(0,self.population_size*6,6):
                individual = self.population[int(i/6)]
                sheet.cell(row = i + 1, column = 1).value = 'rule{}'.format(int(i/6)+1)
                for rule_index in range(len(RULE_SET)):
                    rule_for_X = individual[rule_index]
                    for j in range(DEFAULT_NUM_OF_CONDITION*2):
                        if j == 0 or j == 1:
                            sheet.cell(row = i + 1 + rule_index, column = j + 2).value = OIL_PRICE_LIST[self.convert2to10_in_list(rule_for_X[j])]
                        elif j == 2 or j == 3:
                            sheet.cell(row = i + 1 + rule_index, column = j + 2).value = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule_for_X[j])]
                        else:
                            sheet.cell(row = i + 1 + rule_index, column = j + 2).value = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule_for_X[j])]
                    if RULE_SET[rule_index] == DECISION_SPEED:
                        sheet.cell(row = i + 1 + rule_index, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('change speed to {}'.format(VESSEL_SPEED_LIST[self.convert2to10_in_list(rule_for_X[-1])])
                                                                                                                if self.check_rule_is_adapted(rule_for_X)
                                                                                                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_SELL:
                        sheet.cell(row = i + 1 + rule_index, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('sell {} ships'.format(SELL_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                                                                                                if self.check_rule_is_adapted(rule_for_X)
                                                                                                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_BUY:
                        sheet.cell(row = i + 1 + rule_index, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('buy {} ships'.format(BUY_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                                                                                                if self.check_rule_is_adapted(rule_for_X)
                                                                                                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_CHARTER_OUT:
                        sheet.cell(row = i + 1 + rule_index, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('{0}month charter out, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(rule_for_X[-2])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                                                                                                if self.check_rule_is_adapted(rule_for_X)
                                                                                                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_CHARTER_IN:
                        sheet.cell(row = i + 1 + rule_index, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('{0}month charter in, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(rule_for_X[-2])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                                                                                                if self.check_rule_is_adapted(rule_for_X)
                                                                                                                else 'NOT ADAPTED')
                sheet.cell(row = i + 1 + len(RULE_SET), column = 2).value = individual[-1][0]
                sheet.cell(row = i + 1 + len(RULE_SET), column = 3).value = individual[-1][1]
        else:
            for i in range(0,self.population_size):
                individual = self.population[i]
                sheet.cell(row = i + 1, column = 1).value = 'rule{}'.format(i+1)
                for j in range(DEFAULT_NUM_OF_CONDITION*2):
                    if j == 0 or j == 1:
                        sheet.cell(row = i + 1, column = j + 2).value = OIL_PRICE_LIST[self.convert2to10_in_list(individual[j])]
                    elif j == 2 or j == 3:
                        sheet.cell(row = i + 1, column = j + 2).value = FREIGHT_RATE_LIST[self.convert2to10_in_list(individual[j])]
                    else:
                        sheet.cell(row = i + 1, column = j + 2).value = EXCHANGE_RATE_LIST[self.convert2to10_in_list(individual[j])]
                if self.decision == DECISION_SPEED:
                    sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('change speed to {}'.format(VESSEL_SPEED_LIST[self.convert2to10_in_list(individual[-2])])
                                                                                                if self.check_rule_is_adapted(individual)
                                                                                                else 'NOT ADAPTED')
                elif self.decision == DECISION_SELL:
                    sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('sell {} ships'.format(SELL_NUMBER[self.convert2to10_in_list(individual[-2])])
                                                                                                if self.check_rule_is_adapted(individual)
                                                                                                else 'NOT ADAPTED')
                elif self.decision == DECISION_BUY:
                    sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('buy {} ships'.format(BUY_NUMBER[self.convert2to10_in_list(individual[-2])])
                                                                                                if self.check_rule_is_adapted(individual)
                                                                                                else 'NOT ADAPTED')
                elif self.decision == DECISION_CHARTER_OUT:
                    sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('{0}month charter out, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(individual[-3])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(individual[-2])])
                                                                                                if self.check_rule_is_adapted(individual)
                                                                                                else 'NOT ADAPTED')
                elif self.decision == DECISION_CHARTER_IN:
                    sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2).value = ('{0}month charter in, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(individual[-3])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(individual[-2])])
                                                                                                if self.check_rule_is_adapted(individual)
                                                                                                else 'NOT ADAPTED')
                sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 1 + 2).value = individual[-1][0]
                sheet.cell(row = i + 1, column = DEFAULT_NUM_OF_CONDITION*2 + 2 + 2).value = individual[-1][1]
        wb.save(path)
        wb.close()
        print('saving changes')

    def compare_rules(self):
        fitness_no_rule = self.fitness_function(self.compare_rule)[0]
        if self.decision == DECISION_INTEGRATE:
            name = 'integrate'
            fitness_best = self.bestpopulation[-1][-1][0]
            fitness_full_search = 0
        else:
            if self.decision == DECISION_SPEED:
                name = 'speed'
                fitness_best = self.bestpopulation[-1][-1][0]
                fitness_full_search = self.full_search_method_speed()
            elif self.decision == DECISION_SELL:
                name = 'sell'
                fitness_best = 0
                for i in range(self.population_size):
                    if self.population[i][-2][0] != ACTION_STAY:
                        if self.check_rule_is_adapted(self.population[i]):
                            fitness_best = self.population[i][-1][0]
                            break
                fitness_full_search = self.full_search_method_sell()
            elif self.decision == DECISION_BUY:
                name = 'buy'
                fitness_best = 0
                for i in range(self.population_size):
                    if self.population[i][-2][0] != ACTION_STAY:
                        if self.check_rule_is_adapted(self.population[i]):
                            fitness_best = self.population[i][-1][0]
                            break
                fitness_full_search = 0
            elif self.decision == DECISION_CHARTER_OUT:
                name = 'charter_out'
                fitness_best = 0
                for i in range(self.population_size):
                    if self.population[i][-2][0] != ACTION_STAY:
                        if self.check_rule_is_adapted(self.population[i]):
                            fitness_best = self.population[i][-1][0]
                            break
                fitness_full_search = self.full_search_method_charter()
            elif self.decision == DECISION_CHARTER_IN:
                name = 'charter_out'
                fitness_best = 10
                for i in range(self.population_size):
                    if self.population[i][-2][0] != ACTION_STAY:
                        if self.check_rule_is_adapted(self.population[i]):
                            fitness_best = self.population[i][-1][0]
                            break
                fitness_full_search = 0
            else:
                print('selected decision item does not exist')
                sys.exit()
        print('no rule         ',fitness_no_rule)
        print('best rule       ',fitness_best)
        print('full search rule',fitness_full_search)
        left = [1,2,3]
        height = [fitness_no_rule,fitness_best,fitness_full_search]
        label = ['no rule','best rule','full search']
        plt.title('Comparison among three decision rule')
        plt.ylabel('fitness')
        min_fit = min(fitness_no_rule,min(fitness_best,fitness_full_search))
        max_fit = max(fitness_no_rule,max(fitness_best,fitness_full_search))
        if max_fit < 0:
            plt.ylim(min_fit*1.1,max_fit*0.9)
        else:
            if min_fit < 0:
                plt.ylim(min_fit*1.1,max_fit*1.1)
            else:
                plt.ylim(min_fit*0.9,max_fit*1.1)
        colorlist = ['b','b','b']
        for i in range(len(height)):
            if height[i] < 0:
                colorlist[i] = 'r'
        plt.bar(left,height,color=colorlist,tick_label=label,align='center')
        save_dir = '../output'
        plt.savefig(os.path.join(save_dir, 'comparison_{}.png'.format(name)))
        plt.close()

    def check_rule_is_adapted(self,rule):
        for pattern in range(DEFAULT_PREDICT_PATTERN_NUMBER):
            for year in range(VESSEL_LIFE_TIME):
                for month in range(12):
                    oil_price = self.oil_price_data[pattern][year*12+month]['price']
                    freight= self.freight_rate_outward_data[pattern][year*12+month]['price']
                    exchange = self.exchange_rate_data[pattern][year*12+month]['price']
                    if self.decision == DECISION_INTEGRATE:
                        a = OIL_PRICE_LIST[self.convert2to10_in_list(rule[0])]
                        b = OIL_PRICE_LIST[self.convert2to10_in_list(rule[1])]
                        if a == b or ( a <= oil_price and oil_price <= b):
                            c = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule[2])]
                            d = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule[3])]
                            if c == d or ( c <= freight and freight <= d):
                                e = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule[4])]
                                f = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule[5])]
                                if e == f or ( e <= exchange and exchange <= f):
                                    return True
                    else:
                        result = self.adapt_rule(oil_price,freight,exchange,rule)
                        if result[0]:
                            return True
        return False

    def check_convergence(self,target,criteria):
        flag = True
        for index in range(1,criteria+1):
            if target[-index][-1] != target[-(index+1)][-1]:
                flag = False
                break
        return flag

    def execute_GA(self,multiprocess=None,method=ROULETTE):
        first = time.time()

        #randomly generating individual group
        for i in range(self.population_size):
            self.population.append(self.generateIndividual())
        self.export_excel(0)
        self.depict_average_variance(0,self.population,self.actionlist)

        #genetic algorithm
        for gene in tqdm(range(self.generation)):
            time.sleep(1/self.generation)

            #change population size according to generation
            #self.change_population_size(gene)

            crossing_time = time.time()
            #crossing
            self.temp = copy.deepcopy(self.population)
            for i in range(0,self.population_size,2):
                if random.random() < self.crossover_rate:
                    if self.decision == DECISION_SPEED:
                        a,b = self.crossing(self.temp[i],self.temp[i+1],DEFAULT_NUM_OF_CONDITION*2+1)
                    elif self.decision == DECISION_SELL or self.decision == DECISION_BUY:
                        a,b = self.crossing(self.temp[i],self.temp[i+1],DEFAULT_NUM_OF_CONDITION*2+1)
                    elif self.decision == DECISION_CHARTER_OUT or self.decision == DECISION_CHARTER_IN:
                        a,b = self.crossing(self.temp[i],self.temp[i+1],DEFAULT_NUM_OF_CONDITION*2+2)
                    elif self.decision == DECISION_INTEGRATE:
                        a,b = self.crossing(self.temp[i],self.temp[i+1])
                    else:
                        print('selected decision item does not exist')
                        sys.exit()
                else:
                    a,b = copy.deepcopy(self.temp[i]),copy.deepcopy(self.temp[i+1])
                self.temp.append(a)
                self.temp.append(b)
            print('crossing',time.time()-crossing_time)

            #mutation
            mutation_time = time.time()
            for individual in self.temp:
                if random.random() < self.alpha:
                    individual = self.mutation(individual)
            print('mutation',time.time()-mutation_time)

            #rule check
            self.exchange_rule()

            #computation of fitness
            self.priority = priority
            if multiprocess is None:
                fitness_time = time.time()
                for one in range(len(self.temp)):
                    rule = self.temp[one]
                    rule[-1][0], rule[-1][1] = self.fitness_function(rule,priority)
                print('fitness',time.time()-fitness_time)
            else:
                multifitness_time = time.time()
                num_pool = multi.cpu_count()
                with Pool(num_pool) as pool:
                    p = pool.map(self.multiproces_fitness, range(len(self.temp)))
                    p.sort(key=lambda x:x[0])
                    for i in range(len(p)):
                        self.temp[i][-1][0], self.temp[i][-1][1] = p[i][1]
                print('multifitness',time.time()-multifitness_time)
            #selection
            #change size of self.population
            self.population = [0] * self.population_size
            #store last generation's best individual unchanged
            if gene > 0:
                self.population[0] = self.bestpopulation[gene-1]
            else:#this does not have meaning, just number adjustment
                self.population[0] = self.temp[0]
                #self.depict_average_variance(gene,self.temp)
            if method == ROULETTE:#roulette selection and elite storing
                #store the best 5% individual
                self.temp.sort(key=lambda x:x[-1][0],reverse = True)
                elite_number = int(self.population_size * 0.05)
                for i in range(1,elite_number+1):
                    self.population[i] = self.temp[i]
                min_fit = self.temp[-1][-1][0]
                random.shuffle(self.temp)
                ark = 0 # the number used to roulette in crossing
                probability = 0
                for i in range(len(self.temp)):
                    probability = probability + self.temp[i][-1][0] + (0.1 - min_fit)
                roulette = 0
                for i in range(elite_number+1,self.population_size):
                    roulette = random.randint(0,int(probability))
                    while roulette > 0:
                        roulette = roulette - (self.temp[ark][-1][0] + 0.1 - min_fit)
                        ark = (ark + 1) % len(self.temp)
                    self.population[i] = self.temp[ark]
            elif method == TOURNAMENT:#tournament selection
                for select in range(self.population_size-1):
                    tournament = []
                    for _ in range(6):
                        tournament.append(self.temp[random.randint(0,2*self.population_size-1)])
                    tournament.sort(key=lambda x:x[-1][0],reverse = True)
                    self.population[select] = tournament[0]
            elif method == STEADY_STATE:#steady state ga
                for i in range(0,len(self.temp),2):
                    if i + 1 < self.population_size:
                        store = []
                        store.append(self.temp[i])
                        store.append(self.temp[i+1])
                        if self.population_size + i < len(self.temp):
                            store.append(self.temp[self.population_size + i])
                        if self.population_size + i + 1 < len(self.temp):
                            store.append(self.temp[self.population_size + i+1])
                        store.sort(key=lambda x:x[-1][0],reverse = True)
                        self.population[i] = store[0]
                        self.population[i+1] = store[1]
            else:
                print('Selected method does not exist')
                sys.exit()
            self.population.sort(key=lambda x:x[-1][0],reverse = True)
            self.bestpopulation.append(self.population[0])
            random.shuffle(self.population)
            total = 0
            for e in range(self.population_size):
                total += self.population[e][-1][0]
            self.averagepopulation.append(total/self.population_size)
            #if gene > 10 and self.check_convergence(self.bestpopulation,10):
            #    break

        #print result
        self.population.sort(key=lambda x:x[-1][0],reverse = True)
        for i in range(0,self.population_size):
            if i == 0:
                print('best rule', self.population[i])
            thisone = self.population[i]
            if self.decision == DECISION_INTEGRATE:
                for rule_index in range(len(RULE_SET)):
                    rule_for_X = thisone[rule_index]
                    a = OIL_PRICE_LIST[self.convert2to10_in_list(rule_for_X[0])]
                    b = OIL_PRICE_LIST[self.convert2to10_in_list(rule_for_X[1])]
                    c = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule_for_X[2])]
                    d = FREIGHT_RATE_LIST[self.convert2to10_in_list(rule_for_X[3])]
                    e = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule_for_X[4])]
                    f = EXCHANGE_RATE_LIST[self.convert2to10_in_list(rule_for_X[5])]
                    if RULE_SET[rule_index] == DECISION_SPEED:
                        g = (VESSEL_SPEED_LIST[self.convert2to10_in_list(rule_for_X[-1])]
                                if self.check_rule_is_adapted(rule_for_X)
                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_SELL:
                        g = ('sell {} ships'.format(SELL_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                if self.check_rule_is_adapted(rule_for_X)
                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_BUY:
                        g = ('buy {} ships'.format(BUY_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                if self.check_rule_is_adapted(rule_for_X)
                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_CHARTER_OUT:
                        g = ('{0}month charter out, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(rule_for_X[-2])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                if self.check_rule_is_adapted(rule_for_X)
                                else 'NOT ADAPTED')
                    elif RULE_SET[rule_index] == DECISION_CHARTER_IN:
                        g = ('{0}month charter in, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(rule_for_X[-2])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(rule_for_X[-1])])
                                if self.check_rule_is_adapted(rule_for_X)
                                else 'NOT ADAPTED')
                    if i < NUM_DISPLAY:
                        print('{0} <= oil price <= {1} and {2} <= freight <= {3} and {4} <= exchange <= {5} -> {6}'.format(a,b,c,d,e,f,g))
                if i < NUM_DISPLAY:
                    print('Expectation = {}'.format(thisone[-1][0]))
                    print('Variance = {}'.format(thisone[-1][1]))
            else:
                a = OIL_PRICE_LIST[self.convert2to10_in_list(thisone[0])]
                b = OIL_PRICE_LIST[self.convert2to10_in_list(thisone[1])]
                c = FREIGHT_RATE_LIST[self.convert2to10_in_list(thisone[2])]
                d = FREIGHT_RATE_LIST[self.convert2to10_in_list(thisone[3])]
                e = EXCHANGE_RATE_LIST[self.convert2to10_in_list(thisone[4])]
                f = EXCHANGE_RATE_LIST[self.convert2to10_in_list(thisone[5])]
                if self.decision == DECISION_SPEED:
                    g = ('change speed to {}'.format(VESSEL_SPEED_LIST[self.convert2to10_in_list(thisone[-2])])
                            if self.check_rule_is_adapted(thisone)
                            else 'NOT ADAPTED')
                elif self.decision == DECISION_SELL:
                    g = ('sell {} ships'.format(SELL_NUMBER[self.convert2to10_in_list(thisone[-2])])
                            if self.check_rule_is_adapted(thisone)
                            else 'NOT ADAPTED')
                elif self.decision == DECISION_BUY:
                    g = ('buy {} ships'.format(BUY_NUMBER[self.convert2to10_in_list(thisone[-2])])
                            if self.check_rule_is_adapted(thisone)
                            else 'NOT ADAPTED')
                elif self.decision == DECISION_CHARTER_OUT:
                    g = ('{0}month charter out, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(thisone[-3])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(thisone[-2])])
                            if self.check_rule_is_adapted(thisone)
                            else 'NOT ADAPTED')
                elif self.decision == DECISION_CHARTER_IN:
                    g = ('{0}month charter in, {1} ships'.format(CHARTER_PERIOD[self.convert2to10_in_list(thisone[-3])],CHARTER_SHIPS_NUMBER[self.convert2to10_in_list(thisone[-2])])
                            if self.check_rule_is_adapted(thisone)
                            else 'NOT ADAPTED')
                if i < NUM_DISPLAY and self.check_rule_is_adapted(thisone):
                    print('{0} <= oil price <= {1} and {2} <= freight <= {3} and {4} <= exchange <= {5} -> {6}'.format(a,b,c,d,e,f,g))
                    print('Expectation = {}'.format(thisone[-1][0]))
                    print('Variance = {}'.format(thisone[-1][1]))
            if a > b or c > d or e > f:
                print('rule error')
                sys.exit()
        print('finish')
        print('Spent time is {0}'.format(time.time() - first))
        self.depict_fitness()
        self.depict_average_variance()
        self.export_excel()
        self.compare_rules()

        #initialize attribute
        self.gruop = []
        self.bestpopulation = []
        self.averagepopulation = []

def main():
    generated_sinario = load_generated_sinario()
    oil_data = generated_sinario[0]
    freight_outward_data = generated_sinario[1]
    freight_return_data = generated_sinario[2]
    exchange_data = generated_sinario[3]
    args = sys.argv
    ga = GA(oil_data,freight_outward_data,freight_return_data,exchange_data,
                    TEU_SIZE,INITIAL_SPEED,ROUTE_DISTANCE,
                    [0,0,0,0,0])
    ga.execute_GA()

if __name__ == "__main__":
    main()
